# -*- coding: utf-8 -*-
import io
import base64
from odoo import models, fields, api
from odoo.exceptions import UserError
import unicodedata

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class KhachHang(models.Model):
    _name = 'khach_hang'
    _description = 'Quản lý thông tin khách hàng'
    _rec_name = 'ten_khach_hang'

    ma_khach_hang = fields.Char("Mã khách hàng", readonly=True, copy=False)
    ten_khach_hang = fields.Char("Tên khách hàng", required=True)
    loai_khach_hang = fields.Selection([
        ('ca_nhan', 'Cá nhân'),
        ('doanh_nghiep', 'Doanh nghiệp')
    ], string="Loại khách hàng", default='ca_nhan', required=True)

    email = fields.Char("Email")
    so_dien_thoai = fields.Char("Số điện thoại", required=True)
    dia_chi = fields.Text("Địa chỉ")
    ngay_sinh = fields.Date("Ngày sinh")

    ma_so_thue = fields.Char("Mã số thuế")
    website = fields.Char("Website")

    trang_thai = fields.Selection([
        ('tiem_nang', 'Tiềm năng'),
        ('da_lien_he', 'Đã liên hệ'),
        ('dang_dam_phan', 'Đang đàm phán'),
        ('thanh_cong', 'Thành công'),
        ('that_bai', 'Thất bại')
    ], string="Trạng thái", default='tiem_nang')

    nguoi_phu_trach_id = fields.Many2one('nhan_vien', string="Người phụ trách")
    ghi_chu = fields.Text("Ghi chú")
    phan_hoi_ids = fields.One2many('phan_hoi', 'khach_hang_id', string='Phan hoi')

    # ===== CẢI TIẾN: Thống kê công việc liên quan =====
    so_cong_viec = fields.Integer(
        "Số công việc",
        compute='_compute_so_cong_viec'
    )

    def _compute_so_cong_viec(self):
        CongViec = self.env.get('cong_viec')
        for rec in self:
            if CongViec is not None:
                rec.so_cong_viec = CongViec.search_count([
                    ('khach_hang_id', '=', rec.id)
                ])
            else:
                rec.so_cong_viec = 0

    def action_tao_lich_hen(self):
        """Tạo lịch hẹn → tự động tạo công việc loại lich_hen"""
        for rec in self:
            nhan_vien = rec.nguoi_phu_trach_id
            if not nhan_vien or not nhan_vien.phong_ban_id:
                continue
            self.env['cong_viec'].create({
                'tieu_de': f"Lịch hẹn với {rec.ten_khach_hang}",
                'loai_cong_viec': 'lich_hen',
                'khach_hang_id': rec.id,
                'phong_ban_phu_trach_id': nhan_vien.phong_ban_id.id,
                'nguoi_thuc_hien_id': nhan_vien.id,
                'mo_ta': f"Lịch hẹn tự động tạo cho khách hàng {rec.ten_khach_hang}",
                'muc_do_uu_tien': 'cao',
                'trang_thai': 'moi',
            })
            rec.write({'trang_thai': 'dang_dam_phan'})
        return {
            'type': 'ir.actions.act_window',
            'name': 'Công việc',
            'res_model': 'cong_viec',
            'view_mode': 'list,form',
            'domain': [('khach_hang_id', 'in', self.ids)],
            'context': {'default_khach_hang_id': self.id},
        }

    def action_tao_bao_gia(self):
        """Tạo báo giá → tự động tạo công việc loại gui_bao_gia"""
        for rec in self:
            nhan_vien = rec.nguoi_phu_trach_id
            if not nhan_vien or not nhan_vien.phong_ban_id:
                continue
            self.env['cong_viec'].create({
                'tieu_de': f"Gửi báo giá cho {rec.ten_khach_hang}",
                'loai_cong_viec': 'gui_bao_gia',
                'khach_hang_id': rec.id,
                'phong_ban_phu_trach_id': nhan_vien.phong_ban_id.id,
                'nguoi_thuc_hien_id': nhan_vien.id,
                'mo_ta': f"Báo giá tự động tạo cho khách hàng {rec.ten_khach_hang}",
                'muc_do_uu_tien': 'cao',
                'trang_thai': 'moi',
            })
        return {
            'type': 'ir.actions.act_window',
            'name': 'Công việc',
            'res_model': 'cong_viec',
            'view_mode': 'list,form',
            'domain': [('khach_hang_id', 'in', self.ids)],
            'context': {'default_khach_hang_id': self.id},
        }

    def action_xuat_excel(self):
        """Xuất báo cáo khách hàng ra Excel"""
        if not HAS_OPENPYXL:
            raise UserError("Cần cài openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo cáo khách hàng"

        header_fill = PatternFill("solid", fgColor="6C3483")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["Mã KH", "Tên khách hàng", "Loại KH", "Số điện thoại",
                   "Email", "Trạng thái", "Người phụ trách", "Số công việc", "Ghi chú"]
        col_widths = [12, 25, 15, 15, 25, 18, 20, 12, 30]

        for col, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 25

        trang_thai_map = {
            "tiem_nang": "Tiềm năng", "da_lien_he": "Đã liên hệ",
            "dang_dam_phan": "Đang đàm phán", "thanh_cong": "Thành công", "that_bai": "Thất bại",
        }
        loai_map = {"ca_nhan": "Cá nhân", "doanh_nghiep": "Doanh nghiệp"}
        color_map = {
            "tiem_nang": "D7BDE2", "da_lien_he": "AED6F1",
            "dang_dam_phan": "FAD7A0", "thanh_cong": "A9DFBF", "that_bai": "F1948A",
        }

        records = self.search([]) if not self.ids else self

        for row, rec in enumerate(records, 2):
            row_data = [
                rec.ma_khach_hang or "",
                rec.ten_khach_hang or "",
                loai_map.get(rec.loai_khach_hang, ""),
                rec.so_dien_thoai or "",
                rec.email or "",
                trang_thai_map.get(rec.trang_thai, ""),
                rec.nguoi_phu_trach_id.ho_ten_day_du if rec.nguoi_phu_trach_id else "",
                rec.so_cong_viec,
                rec.ghi_chu or "",
            ]
            fill_color = color_map.get(rec.trang_thai, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_color)

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin
                cell.alignment = Alignment(vertical="center")
                if col == 6:
                    cell.fill = row_fill

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())

        attachment = self.env["ir.attachment"].create({
            "name": "BaoCao_KhachHang.xlsx",
            "type": "binary",
            "datas": file_data,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def action_xem_cong_viec(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Công việc liên quan',
            'res_model': 'cong_viec',
            'view_mode': 'list,form',
            'domain': [('khach_hang_id', '=', self.id)],
            'context': {
                'default_khach_hang_id': self.id,
                'default_nguoi_thuc_hien_id': self.nguoi_phu_trach_id.id,
            },
        }

    def _remove_accents(self, text):
        text = text.replace('Đ', 'D').replace('đ', 'd')
        text = unicodedata.normalize('NFD', text)
        text = text.encode('ascii', 'ignore').decode('utf-8')
        return text

    @api.onchange('ten_khach_hang')
    def _onchange_ma_khach_hang(self):
        if self.ten_khach_hang:
            ten_khong_dau = self._remove_accents(self.ten_khach_hang)
            ma_co_ban = ''.join([word[0].upper() for word in ten_khong_dau.split()])
            ma_co_ban = 'KH' + ma_co_ban
            counter = 1
            ma_khach_hang = f"{ma_co_ban}{counter:02d}"
            while self.search([('ma_khach_hang', '=', ma_khach_hang), ('id', '!=', self.id or 0)], limit=1):
                counter += 1
                ma_khach_hang = f"{ma_co_ban}{counter:02d}"
            self.ma_khach_hang = ma_khach_hang

    @api.model
    def create(self, vals):
        if vals.get('ten_khach_hang'):
            ten_khong_dau = self._remove_accents(vals['ten_khach_hang'])
            ma_co_ban = ''.join([word[0].upper() for word in ten_khong_dau.split()])
            ma_co_ban = 'KH' + ma_co_ban
            counter = 1
            ma_khach_hang = f"{ma_co_ban}{counter:02d}"
            while self.search([('ma_khach_hang', '=', ma_khach_hang)], limit=1):
                counter += 1
                ma_khach_hang = f"{ma_co_ban}{counter:02d}"
            vals['ma_khach_hang'] = ma_khach_hang
        records = super(KhachHang, self).create(vals)
        records._add_birthday_campaign_if_today()
        return records

    def write(self, vals):
        res = super(KhachHang, self).write(vals)
        self._add_birthday_campaign_if_today()
        return res

    def _add_birthday_campaign_if_today(self, today=None):
        today = today or fields.Date.context_today(self)
        if not today:
            return
        birthday_customers = self.filtered(
            lambda c: c.ngay_sinh and (c.ngay_sinh.month, c.ngay_sinh.day) == (today.month, today.day)
        )
        if not birthday_customers:
            return
        Campaign = self.env['chien_dich'].sudo()
        for customer in birthday_customers:
            campaign_name = f"Mừng sinh nhật {customer.ten_khach_hang}"
            campaign = Campaign.search([('ten_chien_dich', '=', campaign_name)], limit=1)
            if not campaign:
                Campaign.create({
                    'ten_chien_dich': campaign_name,
                    'loai_chien_dich': 'khac',
                    'ngay_bat_dau': today,
                    'trang_thai': 'dang_chay',
                    'muc_tieu': f'Chăm sóc sinh nhật cho {customer.ten_khach_hang}.',
                    'khach_hang_ids': [(6, 0, [customer.id])],
                })
            else:
                campaign.write({
                    'ten_chien_dich': campaign_name,
                    'khach_hang_ids': [(4, customer.id)],
                })

    @api.model
    def _cron_add_birthdays_to_campaign(self):
        today = fields.Date.context_today(self)
        if not today:
            return
        birthday_customers = self.search([('ngay_sinh', '!=', False)])
        birthday_customers._add_birthday_campaign_if_today(today=today)